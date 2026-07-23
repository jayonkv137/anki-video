#!/usr/bin/env python3
"""Ingest the German stereotypes library (Excel) → resources/stereotypes_library.json.

Repeatable + idempotent: re-running PRESERVES per-stereotype coverage
(status / episode_id / covered_at) by id, so you never lose "what's been covered"
when the source Excel is re-exported.

Usage: .venv/bin/python scripts/ingest_stereotypes.py [path/to.xlsx]
       (defaults to resources/stereotypes_source.xlsx)
"""
import json
import sys
import datetime
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent
DEFAULT_SRC = REPO / "resources" / "stereotypes_source.xlsx"
OUT = REPO / "resources" / "stereotypes_library.json"


def split_name(name: str) -> tuple[str, str]:
    """'Red Light Enforcement / Bei Rot bleibt man stehen!' → (EN, DE)."""
    if name and " / " in name:
        en, de = name.split(" / ", 1)
        return en.strip(), de.strip()
    return (name or "").strip(), ""


def _s(v) -> str:
    return str(v).strip() if v is not None else ""


def main() -> None:
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_SRC
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)

    # Source references sheet → {index: reference name}
    refs: dict[str, str] = {}
    if "Source references" in wb.sheetnames:
        for r in wb["Source references"].iter_rows(values_only=True):
            if r and r[0] not in (None, "", "Index"):
                refs[_s(r[0])] = _s(r[1]) if len(r) > 1 else ""

    ws = wb["Table 1"] if "Table 1" in wb.sheetnames else wb.worksheets[0]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    rows = [r for r in rows if any(c not in (None, "") for c in r)]

    # Preserve existing coverage across re-ingests
    prev: dict[str, dict] = {}
    if OUT.exists():
        for s in json.loads(OUT.read_text(encoding="utf-8")).get("stereotypes", []):
            prev[s["id"]] = s

    out = []
    for r in rows[1:]:  # skip header
        cat, num, name, desc, ctx, src_key = (list(r) + [None] * 6)[:6]
        sid = _s(num)
        en, de = split_name(_s(name))
        p = prev.get(sid, {})
        out.append({
            "id": sid,
            "category": _s(cat),
            "name": _s(name),
            "name_en": en,
            "name_de": de,
            "description": _s(desc),
            "cultural_context": _s(ctx),
            "source_key": _s(src_key),
            "status": p.get("status", "uncovered"),
            "episode_id": p.get("episode_id"),
            "covered_at": p.get("covered_at"),
        })

    doc = {
        "version": 1,
        "generated_at": datetime.datetime.now().isoformat(timespec="seconds"),
        "source_file": src.name,
        "count": len(out),
        "references": refs,
        "stereotypes": out,
    }
    OUT.write_text(json.dumps(doc, ensure_ascii=False, indent=2), encoding="utf-8")
    covered = sum(1 for s in out if s["status"] == "covered")
    print(f"✓ ingested {len(out)} stereotypes → {OUT.relative_to(REPO)} "
          f"({covered} covered / {len(out) - covered} uncovered); {len(refs)} references")


if __name__ == "__main__":
    main()
